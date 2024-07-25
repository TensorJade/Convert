import pandas as pd
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# 定义正则表达式替换函数
def apply_regex(text, regex_list):
    for pattern, replacement in regex_list:
        text = re.sub(pattern, replacement, text)
    return text

# 定义标准化函数
def standardize_text(text):
    return re.sub(r'\W+', '', text).lower()

# 定义特殊处理函数
def special_rules_match(original_name, modified_name, reference_row, special_rules):
    for rule in special_rules:
        if rule in original_name.lower() and rule not in reference_row['en_name'].lower():
            return False
        if rule not in original_name.lower() and rule in reference_row['en_name'].lower():
            return False

    original_numbers = re.findall(r'\d+', original_name)
    reference_numbers = re.findall(r'\d+', reference_row['en_name'])
    
    if original_numbers and reference_numbers and original_numbers[0] != reference_numbers[0]:
        return False

    return True

# 定义匹配函数
def match_room_types(row, reference_data, threshold=80, regex_list=[], special_rules=[]):
    original_name = row['channel_room_en_name']
    modified_name = apply_regex(original_name, regex_list)
    modified_name = standardize_text(modified_name)
    
    matches = process.extract(modified_name, reference_data['standardized_en_name'], scorer=fuzz.token_sort_ratio, limit=5)
    for match in matches:
        if match[1] >= threshold:
            matched_row = reference_data[reference_data['standardized_en_name'] == match[0]].iloc[0]
            if special_rules_match(original_name, modified_name, matched_row, special_rules):
                return pd.Series([row['room_id'], matched_row['rid'], original_name, matched_row['en_name'], match[1]])
    
    # 当匹配度低于阈值时也输出相关数据
    return pd.Series([row['room_id'], row['rid'], original_name, row['en_name'], match[1]])

# 处理Excel文件
def process_files(input_file, threshold, regex_list, special_rules):
    data = pd.read_excel(input_file).dropna(how='all')
    reference_data = data[['rid', 'en_name']].drop_duplicates()
    reference_data['standardized_en_name'] = reference_data['en_name'].apply(standardize_text)
    
    results = data.apply(match_room_types, axis=1, reference_data=reference_data, threshold=threshold, regex_list=regex_list, special_rules=special_rules)
    results.columns = ['room_id', 'rid', 'channel_room_en_name', 'matched_en_name', 'similarity']
    
    # 获取输出文件名
    output_file = input_file.replace('.xlsx', '_result.xlsx')
    # 保存结果到Excel文件
    results.to_excel(output_file, index=False)
    
    # 高亮未匹配上的行
    wb = load_workbook(output_file)
    ws = wb.active
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="FF0000")

    for row in range(2, ws.max_row + 1):
        # 标记未匹配上的行
        if ws[f'B{row}'].value is None:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill
        else:
            similarity = ws[f'E{row}'].value
            if similarity is not None and similarity < threshold:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
                # 标红不一致的部分
                original_name = ws[f'C{row}'].value
                matched_name = ws[f'D{row}'].value
                if original_name and matched_name:
                    original_numbers = re.findall(r'\d+', original_name)
                    reference_numbers = re.findall(r'\d+', matched_name)
                    if original_numbers and reference_numbers and original_numbers[0] != reference_numbers[0]:
                        ws[f'C{row}'].font = red_font
                        ws[f'D{row}'].font = red_font

    # 保存高亮后的文件
    wb.save(output_file)
    
    messagebox.showinfo("完成", f"匹配完成，结果已保存到 {output_file}")

# UI界面
def create_gui():
    def browse_input_file():
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        input_entry.delete(0, tk.END)
        input_entry.insert(0, filename)

    def add_regex():
        pattern = pattern_entry.get()
        replacement = replacement_entry.get()
        if pattern and replacement:
            regex_list.append((pattern, replacement))
            regex_listbox.insert(tk.END, f"{pattern} -> {replacement}")
            pattern_entry.delete(0, tk.END)
            replacement_entry.delete(0, tk.END)

    def remove_regex():
        selected = regex_listbox.curselection()
        if selected:
            index = selected[0]
            regex_list.pop(index)
            regex_listbox.delete(index)

    def add_special_rule():
        rules = special_rules_text.get("1.0", tk.END).strip().split('\n')
        if rules:
            special_rules.extend(rules)
            special_rules_text.delete("1.0", tk.END)
            special_rules_text.insert(tk.END, "\n".join(special_rules))

    def remove_special_rule():
        current_rules = special_rules_text.get("1.0", tk.END).strip().split('\n')
        for rule in current_rules:
            if rule in special_rules:
                special_rules.remove(rule)
        special_rules_text.delete("1.0", tk.END)
        special_rules_text.insert(tk.END, "\n".join(special_rules))

    def start_processing():
        input_file = input_entry.get()
        try:
            threshold = int(threshold_entry.get())
            if not (0 <= threshold <= 100):
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "阈值应为0到100之间的整数")
            return
        process_files(input_file, threshold, regex_list, special_rules)

    root = tk.Tk()
    root.title("房型匹配工具")
    root.geometry("800x600")
    root.resizable(True, True)

    regex_list = []
    special_rules = ["double", "single", "twin", "executive", "deluxe", "junior", "superior", "premier"]

    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # 输入文件部分
    input_frame = tk.Frame(frame)
    input_frame.pack(fill=tk.X, pady=5)

    tk.Label(input_frame, text="输入文件:").pack(side=tk.LEFT, padx=5)
    input_entry = tk.Entry(input_frame)
    input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    tk.Button(input_frame, text="浏览", command=browse_input_file).pack(side=tk.RIGHT, padx=5)

    # 阈值部分
    threshold_frame = tk.Frame(frame)
    threshold_frame.pack(fill=tk.X, pady=5)

    tk.Label(threshold_frame, text="阈值 (0-100):").pack(side=tk.LEFT, padx=5)
    threshold_entry = tk.Entry(threshold_frame, width=10)
    threshold_entry.pack(side=tk.LEFT, padx=5)
    threshold_entry.insert(tk.END, 0)

    # 正则表达式部分
    regex_frame = tk.Frame(frame)
    regex_frame.pack(fill=tk.BOTH, pady=5, expand=True)

    tk.Label(regex_frame, text="正则表达式模式:").pack(side=tk.LEFT, padx=5)
    pattern_entry = tk.Entry(regex_frame, width=20)
    pattern_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

    tk.Label(regex_frame, text="替换为:").pack(side=tk.LEFT, padx=5)
    replacement_entry = tk.Entry(regex_frame, width=20)
    replacement_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

    tk.Button(regex_frame, text="添加正则表达式", command=add_regex).pack(side=tk.RIGHT, padx=5)
    tk.Button(regex_frame, text="移除选中正则表达式", command=remove_regex).pack(side=tk.RIGHT, padx=5)

    regex_listbox = tk.Listbox(frame, width=80, height=10)
    regex_listbox.pack(fill=tk.BOTH, pady=5, expand=True)

    # 特殊处理规则部分
    special_rules_frame = tk.Frame(frame)
    special_rules_frame.pack(fill=tk.BOTH, pady=5, expand=True)

    tk.Label(special_rules_frame, text="特殊处理规则:").pack(side=tk.TOP, padx=5, anchor='nw')
    special_rules_text = tk.Text(special_rules_frame, width=80, height=10)
    special_rules_text.pack(fill=tk.BOTH, padx=5, pady=5, expand=True)
    special_rules_text.insert(tk.END, "\n".join(special_rules))

    tk.Button(special_rules_frame, text="添加规则", command=add_special_rule).pack(side=tk.LEFT, padx=5, pady=5)
    tk.Button(special_rules_frame, text="移除选中规则", command=remove_special_rule).pack(side=tk.LEFT, padx=5, pady=5)

    # 开始处理按钮
    tk.Button(frame, text="开始处理", command=start_processing).pack(pady=10)

    frame.columnconfigure(1, weight=1)
    frame.rowconfigure(1, weight=1)
    frame.rowconfigure(4, weight=1)
    frame.rowconfigure(7, weight=1)

    root.mainloop()

if __name__ == "__main__":
    create_gui()
